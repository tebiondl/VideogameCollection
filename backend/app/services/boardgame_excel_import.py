from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from difflib import SequenceMatcher
from hashlib import sha256
from io import BytesIO
import json
import math
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy.orm import Session

from .. import models
from .boardgame_player_migration import sync_match_players_from_legacy


MAX_WORKBOOK_BYTES = 10 * 1024 * 1024


def normalize_title(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char)).casefold()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _number(value: Any, minimum: int = 1, maximum: int = 10) -> int | None:
    if value is None or value == "":
        return None
    try:
        parsed = int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return None
    return parsed if minimum <= parsed <= maximum else None


def _rank(value: Any, warnings: list[str], game_name: str) -> int | None:
    if value is None or str(value).strip() in {"", "-"}:
        return None
    try:
        parsed = float(str(value).strip().replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        warnings.append(f"{game_name}: BGG rank '{value}' was not numeric and was left empty.")
        return None
    if not math.isfinite(parsed) or parsed <= 0:
        return None
    if not parsed.is_integer() and parsed < 20:
        converted = int(round(parsed * 1000))
        warnings.append(f"{game_name}: interpreted BGG rank {value} as {converted}.")
        return converted
    return int(round(parsed))


def _price(value: Any, warnings: list[str], game_name: str) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    raw = str(value).strip().replace(",", ".")
    if re.fullmatch(r"\s*\d+(?:\.\d+)?(?:\s*\+\s*\d+(?:\.\d+)?)+\s*", raw):
        total = sum(float(part.strip()) for part in raw.split("+"))
        warnings.append(f"{game_name}: interpreted price '{value}' as €{total:g}.")
        return total
    try:
        return float(raw)
    except ValueError:
        warnings.append(f"{game_name}: price '{value}' was not numeric and was left empty.")
        return None


def _expansions(value: Any) -> str | None:
    raw = _text(value)
    if not raw:
        return None
    values = [item.strip() for item in re.split(r"[,;\n]", raw) if item.strip()]
    return json.dumps(values, ensure_ascii=False) if values else None


def _date_value(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            return None
    raw = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def _header(value: Any) -> str:
    return normalize_title(value).replace(" ", "")


def _find_column(headers: dict[int, str], aliases: set[str]) -> int | None:
    for column, value in headers.items():
        if value in aliases:
            return column
    return None


KNOWN_MATCH_ALIASES = {
    "dune imperium insurreccion": "dune imperium insurrecion",
    "mansiones de la locura": "las mansiones de la locura 2 ed",
    "zombicide 2 ed": "zombicide 2 ed",
    "la tripulacion": "la tripulacioon en busca del noveno planeta",
    "la tripulacion mar profundo": "la tripulacion mision del mar profundo",
    "white castle": "the white castle",
    "tpk": "total party kill",
    "gloomhaven botones y bichos": "gloomhaven botones y bichos",
    "gloomhaven fauces del leon": "gloomhave fauces del leon",
}


def _resolve_owned_name(match_name: str, owned_games: list[dict[str, Any]]) -> dict[str, Any] | None:
    match_key = normalize_title(match_name)
    exact = {normalize_title(game["name"]): game for game in owned_games}
    if match_key in exact:
        return exact[match_key]
    alias_key = KNOWN_MATCH_ALIASES.get(match_key)
    if alias_key and alias_key in exact:
        return exact[alias_key]

    scored = sorted(
        ((SequenceMatcher(None, match_key, key).ratio(), game) for key, game in exact.items()),
        key=lambda item: item[0],
        reverse=True,
    )
    if scored and scored[0][0] >= 0.88:
        return scored[0][1]
    return None


def _game_payload(name: str, section: str, **values: Any) -> dict[str, Any]:
    return {
        "name": name,
        "description": values.get("description"),
        "comments": values.get("comments"),
        "image_url": None,
        "status": "Not Started",
        "mark": values.get("mark"),
        "hype": values.get("hype"),
        "publication_year": None,
        "tags": None,
        "game_type": None,
        "bgg_link": None,
        "library_section": section,
        "bgg_id": None,
        "bgg_rank": values.get("bgg_rank"),
        "price": values.get("price"),
        "expansions": values.get("expansions"),
        "is_expansion": bool(values.get("is_expansion")),
        "parent_game_name": values.get("parent_game_name"),
    }


def parse_boardgame_workbook(content: bytes, filename: str = "workbook.xlsx") -> dict[str, Any]:
    if not content:
        raise ValueError("The workbook is empty.")
    if len(content) > MAX_WORKBOOK_BYTES:
        raise ValueError("The workbook is larger than 10 MB.")
    if not filename.casefold().endswith(".xlsx"):
        raise ValueError("Please upload an .xlsx workbook.")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError("The file is not a readable Excel workbook.") from exc

    games: list[dict[str, Any]] = []
    raw_matches: list[dict[str, Any]] = []
    warnings: list[str] = []
    parsed_blocks: set[tuple[str, int, int]] = set()

    for sheet in workbook.worksheets:
        if sheet.max_row is None or sheet.max_column is None:
            sheet.calculate_dimension(force=True)
        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        for row_index in range(1, min(max_row, 12) + 1):
            row_headers = {
                column: _header(sheet.cell(row_index, column).value)
                for column in range(1, max_column + 1)
                if _header(sheet.cell(row_index, column).value)
            }
            name_columns = [
                column for column, value in row_headers.items()
                if value in {"nombre", "juego", "game", "name"}
            ]
            for name_column in name_columns:
                next_name = min((column for column in name_columns if column > name_column), default=max_column + 1)
                window_end = min(next_name - 1, name_column + 5)
                headers = {column: row_headers.get(column, "") for column in range(name_column, window_end + 1)}
                price_column = _find_column(headers, {"precio", "precioeur", "price"})
                hype_column = _find_column(headers, {"ganas", "anticipacion", "hype"})
                expansion_column = _find_column(headers, {"expansiones", "expansion", "expansions"})
                rating_column = _find_column(headers, {"nota", "rating", "valoracion"})
                rank_column = _find_column(headers, {"bgg", "bggrank", "rankingbgg"})

                section = None
                if price_column or hype_column:
                    section = "wishlist"
                elif expansion_column or rating_column:
                    section = "owned"
                if not section or (sheet.title, row_index, name_column) in parsed_blocks:
                    continue
                parsed_blocks.add((sheet.title, row_index, name_column))

                for data_row in range(row_index + 1, max_row + 1):
                    name = _text(sheet.cell(data_row, name_column).value)
                    if not name:
                        continue
                    rank = _rank(sheet.cell(data_row, rank_column).value, warnings, name) if rank_column else None
                    if section == "wishlist":
                        is_expansion = "expansion" in normalize_title(name)
                        games.append(_game_payload(
                            name,
                            "wishlist",
                            bgg_rank=rank,
                            price=_price(sheet.cell(data_row, price_column).value, warnings, name) if price_column else None,
                            hype=_number(sheet.cell(data_row, hype_column).value) if hype_column else None,
                            is_expansion=is_expansion,
                        ))
                    else:
                        games.append(_game_payload(
                            name,
                            "owned",
                            bgg_rank=rank,
                            mark=_number(sheet.cell(data_row, rating_column).value) if rating_column else None,
                            expansions=_expansions(sheet.cell(data_row, expansion_column).value) if expansion_column else None,
                        ))

            match_game_column = _find_column(row_headers, {"juego", "game", "nombre"})
            players_column = _find_column(row_headers, {"compas", "jugadores", "playedwith", "players"})
            coop_column = _find_column(row_headers, {"coop", "cooperativo", "modo", "mode"})
            outcome_column = _find_column(row_headers, {"victoria", "resultado", "ganador", "outcome", "winner"})
            comments_column = _find_column(row_headers, {"comentarios", "comentario", "comments", "notes"})
            date_column = _find_column(row_headers, {"fecha", "date"})
            if match_game_column and coop_column and outcome_column and (sheet.title, row_index, 0) not in parsed_blocks:
                parsed_blocks.add((sheet.title, row_index, 0))
                occurrence_counter: Counter[str] = Counter()
                for data_row in range(row_index + 1, max_row + 1):
                    game_name = _text(sheet.cell(data_row, match_game_column).value)
                    if not game_name:
                        continue
                    players = _text(sheet.cell(data_row, players_column).value) if players_column else None
                    coop_raw = _text(sheet.cell(data_row, coop_column).value) or ""
                    outcome_raw = _text(sheet.cell(data_row, outcome_column).value) or ""
                    comments = _text(sheet.cell(data_row, comments_column).value) if comments_column else None
                    played_date = _date_value(sheet.cell(data_row, date_column).value) if date_column else None

                    coop_key = normalize_title(coop_raw)
                    outcome_key = normalize_title(outcome_raw)
                    if coop_key in {"si", "yes", "coop", "cooperative", "cooperativo"}:
                        mode = "cooperative" if players else "solo"
                    else:
                        mode = "competitive" if players else "solo"

                    winner_name = None
                    if outcome_key in {"no terminado", "no terminada", "unfinished", "incomplete"}:
                        result = "incomplete"
                    elif mode in {"cooperative", "solo"}:
                        result = "victory" if outcome_key in {"victoria", "victory", "win"} else "defeat" if outcome_key in {"derrota", "defeat", "loss"} else "incomplete"
                    else:
                        result = "winner"
                        if outcome_key in {"derrota", "defeat", "loss"}:
                            winner_name = "Opponent (not specified)"
                            warnings.append(f"{game_name} row {data_row}: competitive defeat had no winner name.")
                        elif outcome_key in {"victoria", "victory", "win"}:
                            winner_name = "Me"
                        else:
                            winner_name = outcome_raw or "Unknown"

                    fingerprint_data = [game_name, players, mode, result, winner_name, comments, played_date]
                    fingerprint = json.dumps(fingerprint_data, ensure_ascii=False, separators=(",", ":"))
                    occurrence_counter[fingerprint] += 1
                    import_key = sha256(f"boardgame-excel-v1:{fingerprint}:{occurrence_counter[fingerprint]}".encode("utf-8")).hexdigest()
                    raw_matches.append({
                        "game_name": game_name,
                        "played_with": json.dumps([item.strip() for item in re.split(r"[,;]", players) if item.strip()], ensure_ascii=False) if players else None,
                        "mode": mode,
                        "result": result,
                        "winner_name": winner_name,
                        "comments": comments,
                        "played_date": played_date,
                        "import_key": import_key,
                        "source_row": data_row,
                    })

    wishlist_games = [game for game in games if game["library_section"] == "wishlist"]
    owned_games = [game for game in games if game["library_section"] == "owned"]
    if not wishlist_games and not owned_games and not raw_matches:
        raise ValueError("No recognizable board-game tables were found in the workbook.")

    match_only_names: list[str] = []
    for match in raw_matches:
        resolved = _resolve_owned_name(match["game_name"], owned_games)
        if not resolved:
            resolved = _game_payload(
                match["game_name"],
                "owned",
                comments="Added automatically from imported match history.",
            )
            owned_games.append(resolved)
            games.append(resolved)
            match_only_names.append(resolved["name"])
        match["resolved_game_name"] = resolved["name"]

    unknown_dates = sum(match["played_date"] is None for match in raw_matches)
    incomplete_matches = sum(match["result"] == "incomplete" for match in raw_matches)
    if unknown_dates:
        warnings.append(f"{unknown_dates} matches have no date and will be shown as 'Date unknown'.")
    if incomplete_matches:
        warnings.append(f"{incomplete_matches} matches are marked as not finished.")
    if match_only_names:
        warnings.append(f"{len(match_only_names)} owned-game entries will be created from match history so every match has a game.")

    return {
        "filename": filename,
        "games": games,
        "matches": raw_matches,
        "warnings": warnings,
        "stats": {
            "wishlist_games": len(wishlist_games),
            "owned_games": len(owned_games),
            "match_only_games": len(match_only_names),
            "matches": len(raw_matches),
            "unknown_dates": unknown_dates,
            "incomplete_matches": incomplete_matches,
        },
        "samples": {
            "wishlist": [game["name"] for game in wishlist_games[:8]],
            "owned": [game["name"] for game in owned_games[:8]],
            "match_games": list(dict.fromkeys(match["resolved_game_name"] for match in raw_matches))[:8],
        },
    }


def build_import_preview(content: bytes, filename: str, db: Session, user_id: int) -> dict[str, Any]:
    parsed = parse_boardgame_workbook(content, filename)
    existing_game_keys = {
        (normalize_title(game.name), game.library_section)
        for game in db.query(models.Boardgame).filter(models.Boardgame.user_id == user_id).all()
    }
    existing_import_keys = {
        value for (value,) in db.query(models.BoardgameMatch.import_key).filter(
            models.BoardgameMatch.user_id == user_id,
            models.BoardgameMatch.import_key.isnot(None),
        ).all()
    }
    parsed["stats"]["games_already_present"] = sum(
        (normalize_title(game["name"]), game["library_section"]) in existing_game_keys
        for game in parsed["games"]
    )
    parsed["stats"]["matches_already_present"] = sum(
        match["import_key"] in existing_import_keys for match in parsed["matches"]
    )
    return {key: parsed[key] for key in ("filename", "warnings", "stats", "samples")}


def commit_boardgame_import(content: bytes, filename: str, db: Session, user_id: int) -> dict[str, Any]:
    parsed = parse_boardgame_workbook(content, filename)
    existing_games = db.query(models.Boardgame).filter(models.Boardgame.user_id == user_id).all()
    game_map = {(normalize_title(game.name), game.library_section): game for game in existing_games}
    created_games = 0
    updated_games = 0

    imported_fields = ("bgg_rank", "price", "hype", "mark", "expansions", "is_expansion", "parent_game_name")
    try:
        for payload in parsed["games"]:
            key = (normalize_title(payload["name"]), payload["library_section"])
            game = game_map.get(key)
            if game:
                changed = False
                for field in imported_fields:
                    value = payload.get(field)
                    if value is not None and getattr(game, field) != value:
                        setattr(game, field, value)
                        changed = True
                if changed:
                    updated_games += 1
                continue
            game = models.Boardgame(**payload, user_id=user_id)
            db.add(game)
            db.flush()
            game_map[key] = game
            created_games += 1

        existing_import_keys = {
            value for (value,) in db.query(models.BoardgameMatch.import_key).filter(
                models.BoardgameMatch.user_id == user_id,
                models.BoardgameMatch.import_key.isnot(None),
            ).all()
        }
        created_matches = 0
        skipped_matches = 0
        for payload in parsed["matches"]:
            if payload["import_key"] in existing_import_keys:
                skipped_matches += 1
                continue
            game_key = (normalize_title(payload["resolved_game_name"]), "owned")
            game = game_map.get(game_key)
            if not game:
                raise ValueError(f"Could not resolve imported game: {payload['resolved_game_name']}")
            match = models.BoardgameMatch(
                user_id=user_id,
                boardgame_id=game.id,
                played_with=payload["played_with"],
                mode=payload["mode"],
                result=payload["result"],
                winner_name=payload["winner_name"],
                comments=payload["comments"],
                # Existing production databases created this column as NOT NULL.
                # An empty string preserves an unknown legacy date without inventing one.
                played_date=payload["played_date"] or "",
                import_key=payload["import_key"],
            )
            db.add(match)
            db.flush()
            sync_match_players_from_legacy(db, match)
            existing_import_keys.add(payload["import_key"])
            created_matches += 1
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "filename": filename,
        "games_created": created_games,
        "games_updated": updated_games,
        "games_skipped": len(parsed["games"]) - created_games - updated_games,
        "matches_created": created_matches,
        "matches_skipped": skipped_matches,
        "warnings": parsed["warnings"],
    }
