from io import BytesIO
from thefuzz import process
from .models import Game
from typing import List, Dict, Any, Optional
import openpyxl

# Map internal DB columns to potential human-readable headers (Spanish/English)
COLUMN_MAPPING_TARGETS = {
    "title": ["Title", "Título", "Juego", "Game", "Nombre"],
    "status": ["Status", "Estado", "Lista", "List"],
    "hype_score": ["Hype", "Ganas", "Score", "Puntuación"],
    "rating": ["Rating", "Nota", "Score", "Puntuación"],
    "progress": ["Progress", "Progreso", "Estado Juego"],
    "playtime_hours": ["Playtime", "Horas", "Tiempo", "Hours", "Duration"],
    "finish_year": ["Finish Year", "Año Terminado", "Terminado", "Finished"],
    "release_year": ["Release Year", "Año Lanzamiento", "Lanzamiento", "Released"],
    "price": ["Price", "Precio", "Coste"],
    "platform": ["Platform", "Plataforma", "Consola", "System"],
    "steam_deck": ["Steam Deck", "Deck", "Portable"],
    "notes": ["Notes", "Notas", "Comentarios"],
}


def parse_excel_file(file_content: bytes) -> Dict[str, List[Dict[str, Any]]]:
    """
    Parses an Excel file using openpyxl to extract values AND background colors.
    Returns a dict where keys are sheet names and values are list of records.
    Each record is: { "Header": { "v": value, "c": "FFFFFF" } }
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows())
        if not rows:
            continue

        # Get headers from first row
        headers = []
        for cell in rows[0]:
            h_val = (
                str(cell.value).strip()
                if cell.value is not None
                else f"Unnamed:{cell.column}"
            )
            headers.append(h_val)

        records = []
        for row in rows[1:]:
            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            row_data = {}
            for i, cell in enumerate(row):
                if i >= len(headers):
                    break

                val = cell.value

                # Extract Color
                color = None
                if cell.fill and cell.fill.patternType == "solid":
                    raw_color = cell.fill.start_color
                    if raw_color.type == "rgb":
                        # ARGB hex string
                        if raw_color.rgb and len(raw_color.rgb) >= 6:
                            # Keep last 6 chars (RGB)
                            color = raw_color.rgb[-6:]
                    # Note: Theme colors are complex to resolve, skipping for now
                    # user likely uses standard colors if they color-code manually

                row_data[headers[i]] = {"v": val, "c": color}
            records.append(row_data)

        if records:
            result[sheet_name] = records

    return result


def propose_mapping(headers: List[str]) -> Dict[str, Dict[str, Any]]:
    """
    For each DB column, find the best matching header from the file.
    Returns:
    {
        "db_col_name": {
            "best_match": "Header Name",
            "score": 90,
            "alternatives": ["Other Header", "Another"]
        }
    }
    """
    mapping = {}

    for db_col, candidates in COLUMN_MAPPING_TARGETS.items():
        best_match = None
        best_score = 0
        alternatives = []  # type: List[str]

        header_scores = []
        for header in headers:
            # Check if header matches any of the candidates
            extract = process.extractOne(str(header), candidates)
            if extract:
                score = extract[1]
                header_scores.append((header, score))

        # Sort headers by score descending
        header_scores.sort(key=lambda x: x[1], reverse=True)

        if header_scores:
            best_match = header_scores[0][0]
            best_score = header_scores[0][1]
            alternatives = [h[0] for h in header_scores[1:3]]

            mapping[db_col] = {
                "selected": best_match if best_score > 60 else None,
                "score": best_score,
                "alternatives": alternatives,
            }

    return mapping


def fuzzy_find_game(title: str, existing_games: List[Game]) -> Game:
    """
    Finds an existing game in the user's library that matches the title.
    """
    if not title or not existing_games:
        return None

    choices = {g.title: g for g in existing_games}
    extract = process.extractOne(title, choices.keys())

    if extract and extract[1] >= 90:  # High threshold for automatic matching
        return choices[extract[0]]

    return None
